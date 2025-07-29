import pyautogui
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage # Import per inserire immagini
import os
import traceback
from PIL import Image

# Percorsi delle immagini dei pulsanti OK dei popup
OK_BUTTON_IMAGE_PATH_1 = r'C:\Users\s.bonfanti\Lavori stefano\automations\attivita_giornaliere_project\ok_button.png'
LOGO_PATH = r'C:\Users\s.bonfanti\Lavori stefano\automations\attivita_giornaliere_project\logo.png'

def handle_popup(image_paths):
    """Cerca i pulsanti OK dei popup e li clicca se presenti."""
    for _ in range(5):
        for img_path in image_paths:
            try:
                if pyautogui.locateOnScreen(img_path, confidence=0.8, grayscale=True):
                    print(f"Popup rilevato con {img_path}! Premo Invio...")
                    pyautogui.press('enter')
                    time.sleep(1)
                    return True
            except Exception:
                pass # Ignora eccezioni e continua
        time.sleep(0.5)
    return False

# --- Definizione Costanti ---
COLUMNS_TO_DROP_BY_NAME = ['Cod. attività', 'Dt. ins', 'Dt. scad.', 'Classe att.', 'Stato', 'Priorità', 'Responsabile', 'Titolare', 'Tipo att.', 'Prodotto', 'Destinazione', 'Nr. lotto', 'Macchina', 'Cod. fase', 'Dt. chiusura', 'Gravità', 'Nr. rifer.', 'Dt. rifer.', 'Tipo ass.', 'Modalità ass.', 'Nr. Rintracciabilità', 'Tipo fatt.', 'Situaz. fatt.']
ANALYSIS_COLUMN_NAME = "Inseritore"
SUBJECT_COLUMN_NAME = "Oggetto e descrizione"
DATE_COLUMN_NAME = "Dt. ins."
NOTE_INTERNE_COLUMN_NAME = "Note interne 1"

def classify_soggetto(soggetto):
    """Classifica il soggetto in base a parole chiave specifiche."""
    soggetto_str = str(soggetto).lower()
    if 'contatto cliente' in soggetto_str:
        return 'Contatto Cliente'
    return 'Azione Commerciale'

def process_excel_file(input_file_path, start_date, end_date):
    """
    Elabora il file Excel: filtra per data, elimina colonne, crea fogli formattati e rimuove il foglio originale.
    """
    print(f"Inizio elaborazione del file Excel: {input_file_path} per il periodo dal {start_date} al {end_date}")
    try:
        df = pd.read_excel(input_file_path)

        df[DATE_COLUMN_NAME] = pd.to_datetime(df[DATE_COLUMN_NAME], errors='coerce').dt.date
        df_filtered = df[(df[DATE_COLUMN_NAME] >= start_date) & (df[DATE_COLUMN_NAME] <= end_date)].copy()

        if df_filtered.empty:
            print(f"Nessun dato trovato per il periodo dal {start_date} al {end_date}.")
            return

        df_processed = df_filtered.drop(columns=COLUMNS_TO_DROP_BY_NAME, errors='ignore')
        
        inseritore_replacements = {
            'ab001': 'Alessandra', 'gmoro': 'Gabriella', 'martines': 'Martine',
            'rpacini': 'Rachele', 'r.saber': 'Rachida', 'frosi': 'Federico'
        }
        if ANALYSIS_COLUMN_NAME in df_processed.columns:
            df_processed[ANALYSIS_COLUMN_NAME] = df_processed[ANALYSIS_COLUMN_NAME].str.lower().replace(inseritore_replacements)
        else:
            print(f"Errore: Colonna '{ANALYSIS_COLUMN_NAME}' non trovata.")
            return

        if SUBJECT_COLUMN_NAME not in df_processed.columns:
            print(f"Errore: Colonna '{SUBJECT_COLUMN_NAME}' non trovata, impossibile categorizzare.")
            return

        df_processed['Categoria'] = df_processed[SUBJECT_COLUMN_NAME].apply(classify_soggetto)
        unique_inseritori = df_processed[ANALYSIS_COLUMN_NAME].unique()

        # Creazione di un nuovo ExcelWriter in modalità scrittura per sovrascrivere il file originale
        with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='w') as writer:
            for inseritore in unique_inseritori:
                df_inseritore = df_processed[df_processed[ANALYSIS_COLUMN_NAME] == inseritore]
                unique_categories = df_inseritore['Categoria'].unique()

                for categoria in unique_categories:
                    df_categoria = df_inseritore[df_inseritore['Categoria'] == categoria]
                    df_to_write = df_categoria.drop(columns=[ANALYSIS_COLUMN_NAME, SUBJECT_COLUMN_NAME, 'Categoria'], errors='ignore')
                    
                    safe_inseritore = "".join(c for c in str(inseritore) if c.isalnum() or c in (' ', '_')).rstrip()
                    safe_categoria = "".join(c for c in str(categoria) if c.isalnum() or c in (' ', '_')).rstrip()
                    sheet_name = f"{safe_inseritore}_{safe_categoria}"[:31]

                    df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"Creato foglio per: {inseritore} - {categoria}")

                    worksheet = writer.sheets[sheet_name]
                    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
                    worksheet.page_setup.fitToPage = True
                    worksheet.page_setup.fitToWidth = 1
                    worksheet.page_setup.fitToHeight = 0
                    worksheet.print_options.print_grid_lines = True
                    worksheet.print_options.print_headings = True
                    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
                    worksheet.page_setup.horizontalCentered = True
                    worksheet.page_setup.verticalCentered = True

                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    header_font = Font(bold=True, size=13)
                    default_font = Font(size=14)

                    # Inserisce due righe per il logo e il titolo
                    worksheet.insert_rows(1)
                    worksheet.insert_rows(1)

                    # Inserisce il logo in A1
                    try:
                        img = OpenpyxlImage(LOGO_PATH)
                        # Puoi ridimensionare l'immagine se necessario, ad esempio:
                        # img.width = 100
                        # img.height = 100
                        worksheet.add_image(img, 'A1')
                    except Exception as e:
                        print(f"Errore nell'inserimento del logo: {e}")

                    # Il titolo ora va in A2
                    title = f"{inseritore} - {categoria}"
                    worksheet['A2'] = title
                    worksheet['A2'].font = Font(bold=True, size=28)
                    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=df_to_write.shape[1])
                    worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')

                    MAX_COLUMN_WIDTH = 30
                    for col_idx, col in enumerate(worksheet.columns, 1):
                        max_length = 0
                        column = get_column_letter(col_idx)
                        for cell_idx, cell in enumerate(col, 1):
                            # Le righe sono spostate in basso di 2 (logo in riga 1, titolo in riga 2)
                            if cell_idx > 2: # Salta la riga del logo e del titolo grande
                                cell.border = thin_border
                                if cell_idx == 3: # Intestazioni (ora riga 3)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                                else: # Dati (ora riga 4 in poi)
                                    cell.font = default_font
                                    cell.alignment = Alignment(wrapText=True)
                                
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                        
                        header_text = worksheet.cell(row=3, column=col_idx).value # Le intestazioni sono ora in riga 3
                        if header_text == NOTE_INTERNE_COLUMN_NAME:
                            adjusted_width = 80
                        else:
                            adjusted_width = min((max_length + 4), MAX_COLUMN_WIDTH)
                        worksheet.column_dimensions[column].width = adjusted_width
        
        print("Elaborazione Excel completata con successo!")

    except FileNotFoundError:
        print(f"Errore: File '{input_file_path}' non trovato.")
    except Exception as e:
        print(f"Errore imprevisto durante l'elaborazione del file: {e}")
        traceback.print_exc()

def run():
    analysis_choice = input("Per quale data vuoi analizzare i dati? (digita 'oggi', 'settimana corrente' o una data specifica): ").lower()

    start_date = None
    end_date = None

    if analysis_choice == 'oggi':
        start_date = datetime.date.today()
        end_date = datetime.date.today()
        analysis_filename_date_str = start_date.strftime("%Y_%m_%d")
    elif analysis_choice == 'settimana corrente':
        today = datetime.date.today()
        # Calcola il lunedì della settimana corrente (Monday is 0, Sunday is 6)
        start_date = today - datetime.timedelta(days=today.weekday())
        # Calcola la domenica della settimana corrente
        end_date = start_date + datetime.timedelta(days=6)
        analysis_filename_date_str = f"{start_date.strftime("%Y_%m_%d")}_to_{end_date.strftime("%Y_%m_%d")}"
    else:
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d-%m-%y', '%d/%m/%y'):
            try:
                specific_date = datetime.datetime.strptime(analysis_choice, fmt).date()
                start_date = specific_date
                end_date = specific_date
                analysis_filename_date_str = start_date.strftime("%Y_%m_%d")
                break
            except ValueError:
                pass
    
    if not start_date:
        print("Formato data non valido. Uscita.")
        return

    print("Avvio automazione per generare il report aggiornato...")
    # --- Blocco Automazione PyAutoGUI ---
    pyautogui.hotkey('win', 'r')
    time.sleep(1)
    pyautogui.write(r'C:\Program Files (x86)\ProdWare\Bin\prodware.exe')
    pyautogui.press('enter')
    time.sleep(8)
    handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=960, y=580); time.sleep(8); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=1739, y=344); time.sleep(8); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=90, y=85); time.sleep(8); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=819, y=104); time.sleep(8); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.rightClick(x=9, y=193); time.sleep(3); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=56, y=264); time.sleep(3); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    pyautogui.click(x=765, y=525); time.sleep(3); handle_popup([OK_BUTTON_IMAGE_PATH_1])
    time.sleep(5)
    
    # Il nome del file Excel ora riflette la data o l'intervallo di date di analisi
    full_path = os.path.join(os.getcwd(), f"OpzioniEsportazione_{analysis_filename_date_str}.xlsx")
    
    pyautogui.write(full_path); time.sleep(1); pyautogui.press('enter'); time.sleep(3)
    
    pyautogui.click(x=1045, y=568); time.sleep(3)
    pyautogui.click(x=1100, y=566); time.sleep(3)

    print("Clicco sulle coordinate fisse del popup di conferma: (1081, 588)")
    pyautogui.click(x=1081, y=588); time.sleep(5)

    pyautogui.click(x=1895, y=8); time.sleep(2)
    pyautogui.click(x=1896, y=10); time.sleep(2)
    print("Automazione completata.")
    # --- Fine Blocco Automazione ---

    max_wait_time = 60
    start_time_wait = time.time()
    print(f"Tentativo di elaborazione del file Excel: {full_path}")
    while time.time() - start_time_wait < max_wait_time:
        if os.path.exists(full_path):
            try:
                process_excel_file(full_path, start_date, end_date)
                print("Script completato con successo.")
                return
            except (PermissionError, OSError) as e:
                print(f"File temporaneamente bloccato: {e}. Riprovo...")
                time.sleep(2)
            except Exception as e:
                print(f"Errore inatteso durante l'elaborazione: {e}")
                traceback.print_exc()
                break
        else:
            time.sleep(2)

    print(f"Errore: Impossibile elaborare il file Excel dopo {max_wait_time} secondi: {full_path}")

if __name__ == "__main__":
    run()